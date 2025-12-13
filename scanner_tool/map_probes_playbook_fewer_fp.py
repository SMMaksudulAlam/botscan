import openpyxl
import json
import pickle

length = 16

def filter(r):
	d = {}
	flag_c2 = 1
	flag_m = 1
	for i in range(len(r)):
		fInd = r[i].find(':')
		if(r[i][:fInd+1] != r[i]):
			if(r[i][:fInd] == "c2_1"):
				if(flag_c2 == 1):
					d[r[i]] = 1
					flag_c2 = 0
				else:
					if(d.get(r[i])!=None):
						r = r[i:]
						return filter(r)
			if(r[i][:fInd] == "m_1"):
				if(flag_m == 1):
					d[r[i]] = 1
					flag_m = 0
				else:
					if(d.get(r[i])!=None):
						r = r[i:]
						r.insert(0, 'c2_:')
						return filter(r)			
	return r

def filter_(r):
	
	for i in range(len(r)):
		fInd_1 = r[i].find('_')
		#fInd_1_ = r[i].find('@')
		fInd_2 = r[i].find(':')
		if(r[i][0] == "m"):
			r[i] = r[i][:fInd_1]+r[i][fInd_2:]
		else:
			r[i] = r[i][:fInd_1]+r[i][fInd_2:]
	return r

def map_probes_(dic, r):
	r = filter(r)
	r = filter_(r)
	temp = dic
	for i in range(len(r)):
		if(r[i] not in temp.keys()):
			temp[r[i]] = {}
		temp = temp[r[i]]

def write_to_excel(data, filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(filename)


def map_probes(sheet):
	global length
	data = []
	for row in sheet.iter_rows(min_row=6, values_only=True): 
		row = list(row)
		row.pop(0)
		c2_sig = None
		m_sig = None
		if(row[0]==None):
			continue
		turn = 1
		ind = 4
		temp = []
		while(True):
			if(row[ind] == None or (row[ind][:2] != "c2" and row[ind][:2] != "m_")):
				break
			if(turn == 1):
				if(row[ind][:2]=="c2"):
					temp.append(row[ind])
					ind+=1
				else:
					temp.append("c2:")
				turn = 0
			elif(turn == 0):
				if(row[ind][:2]=="m_"):
					temp.append(row[ind])
					ind+=1
				else:
					temp.append("m:")
				turn = 1
			else:
				continue
		data.append(temp)
	write_to_excel(data, "modified.xlsx")
	dic = {}
	#print(data)
	first_probe = []
	map_fProbe = {}
	for i in range(len(data)):
		r = filter(data[i])
		r = filter_(r)
		flag = 0
		fProbe = []
		if(r and r[-1][0]!='m'):
			r.append('m:')
		print(">>>", r)
		for j in range(min(len(r), 20)):
			if(r[j] == None):
				break
			if(r[j][0] == 'm'):
				continue
			if(r[j][0]=='c'):
				if(flag == 0):
					if(r[j][3:]==""):
						if(r[j+1][0]=='m'):
							x = r[j+1][2:]
							x = x.replace('31302e302e322e3135', '33342e3134352e32332e323238') #changing IPs for tricky ones
							fProbe.append(x)
							#fPro += r[j+1][2:]
							#print("\n\n changed ip \n\n")
					else:
						"""
						if(map_fProbe.get(fPro) == None and fPro != ""):
							first_probe.append(fProbe)
						"""
						fPro = "".join(fProbe)
						map_fProbe[fPro] = map_fProbe.get(fPro, 0) + 1
						print(f"{i} {fPro} : {map_fProbe[fPro]}")
						if(map_fProbe[fPro]==4 and fPro!=""):
							first_probe.append(fProbe)
						flag = 1
				if(flag == 1):
					if(dic.get(r[j][3:]) == None and r[j+1][0]=='m'):
						dic[r[j][3:]] = {}
					curD = dic[r[j][3:]]
					if(r[j+1][0]=='m'):
						if(curD.get(r[j+1][2:]) == None):
							curD[r[j+1][2:]] = 1
						else:
							curD[r[j+1][2:]] += 1
							
							
	print(map_fProbe)
	with open('first_probe.pkl', 'wb') as file:
		print(first_probe)
		pickle.dump(first_probe, file)

	for k in dic.keys():
		dic[k] = dict(sorted(dic[k].items(), key=lambda item: item[1], reverse=True))


	
	benign_payloads = ["5353482d322e302d4f70656e5353485f382e3270", "5353482d322e302d4f70656e5353485f382e3070", "485454502f312e3120323030204f4b0d0a446174", "250a"]
	
	need_to_delete = []
	for k in dic.keys():
		sub_key = k[:40]
		if(sub_key in benign_payloads):
			need_to_delete.append(k)
	for k in need_to_delete:
		del dic[k]


	with open('mapped_probes.json', 'w') as file:
		json.dump(dic, file, indent=4)




workbook = openpyxl.load_workbook('playbook.xlsx')

sheet = workbook['signatures'] 
map_probes(sheet)


